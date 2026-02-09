"""Extract text from Excel, one chunk per sheet."""

from pathlib import Path
from app.domain.documents import Chunk


def extract_xlsx(filepath: str) -> list[Chunk]:
    from openpyxl import load_workbook

    wb = load_workbook(filepath, data_only=True)
    fname = Path(filepath).name
    chunks: list[Chunk] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            line = " | ".join(cells)
            if line.replace("|", "").strip():
                rows.append(line)

        if rows:
            chunks.append(Chunk(
                source_file=fname, source_type="xlsx",
                section_title=sheet_name, content="\n".join(rows),
            ))

    return chunks
